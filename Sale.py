from flask import Flask, render_template, request
import pandas as pd
import os
from openpyxl import load_workbook

app = Flask(__name__)

FILE = os.path.join(app.root_path, "Computers.xlsx")


@app.route('/', methods=['GET', 'POST'])
def index():
    message = ""

    
    try:
        df_devices = pd.read_excel(FILE, sheet_name="Sheet1")
        df_devices.columns = df_devices.columns.str.strip()
    except Exception as e:
        return f"❌ Error reading Sheet1: {e}"


    required_cols = ["Device", "Model", "Ram", "Processor", "Comments"]
    for col in required_cols:
        if col not in df_devices.columns:
            df_devices[col] = ""


    try:
        df_sales = pd.read_excel(FILE, sheet_name="Madesales")
        df_sales.columns = df_sales.columns.str.strip()
    except:
        df_sales = pd.DataFrame(columns=["device", "employee_name", "employee_id"])

    assigned_list = df_sales["device"].astype(str).tolist()

    df_available = df_devices[
        ~df_devices["Device"].astype(str).isin(assigned_list)
    ].copy()

    df_available = df_available.fillna("")
    df_available["display"] = (
        df_available["Device"].astype(str) + "  | " +
        df_available["Model"].astype(str) + "   | " +
        df_available["Ram"].astype(str) + "     | " +
        df_available["Processor"].astype(str) + "   | " +
        df_available["Comments"].astype(str)
    )

    computers = df_available.to_dict("records")

 
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        emp_id = request.form.get("emp_id", "").strip()
        device = request.form.get("device", "").strip()


        try:
            df_sales_latest = pd.read_excel(FILE, sheet_name="Madesales")
            df_sales_latest.columns = df_sales_latest.columns.str.strip()
        except:
            df_sales_latest = pd.DataFrame(columns=["device", "employee_name", "employee_id"])
     
        if not emp_id.isdigit():
            message="Please use only numbers"
            return render_template("SALE.html",computers=computers,message=message)
    
        if device in df_sales_latest["device"].astype(str).values:
            message = "❌ Someone else already took this device. Please choose another."
            return render_template("SALE.html", computers=computers, message=message)

       
        if emp_id in df_sales_latest["employee_id"].astype(str).values:
            message = "❌ This employee already has a device!"
            return render_template("SALE.html", computers=computers, message=message)

     
        new_record = {
            "device": device,
            "employee_name": name,
            "employee_id": emp_id
        }
        df_sales_latest = pd.concat([df_sales_latest, pd.DataFrame([new_record])], ignore_index=True)

        try:
            wb = load_workbook(FILE)

           
            if "Madesales" in wb.sheetnames:
                wb.remove(wb["Madesales"])
            wb.save(FILE)

   
            with pd.ExcelWriter(FILE, engine="openpyxl", mode="a") as writer:
                df_sales_latest.to_excel(writer, sheet_name="Madesales", index=False)

            message = "✅ Saved successfully!"

        except PermissionError:
            message = "❌ Please CLOSE Computers.xlsx first!"
        except Exception as e:
            message = f"❌ Error saving: {e}"

    return render_template("SALE.html", computers=computers, message=message)


if __name__ == '__main__':
    app.run(debug=True)